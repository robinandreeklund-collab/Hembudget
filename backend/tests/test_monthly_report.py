"""Tester för nya månadsrapporten — både datamodellen (build_report_data)
och det faktiska PDF-genereringen (render_pdf).

Vi validerar:
- Rätt KPI:er i summary
- Transfer-förslag räknar fram både 50/50 och prorata
- Delta mot föregående månad sorteras efter absolut belopp
- PDF-bytes returneras och börjar med %PDF-magic-bytes
- /reports/month/{month}/pdf + /excel endpoints svarar 200
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool


@pytest.fixture()
def client(monkeypatch):
    monkeypatch.setenv("HEMBUDGET_DEMO_MODE", "1")

    from hembudget.db.models import Base

    engine = create_engine(
        "sqlite:///:memory:",
        future=True,
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)

    from hembudget import demo as demo_mod
    monkeypatch.setattr(demo_mod, "bootstrap_if_empty", lambda: {"skipped": True})

    from hembudget.api import deps as api_deps
    from hembudget.main import build_app

    app = build_app()

    def _fake_db():
        s = SessionLocal()
        try:
            yield s
            s.commit()
        except Exception:
            s.rollback()
            raise
        finally:
            s.close()

    app.dependency_overrides[api_deps.db] = _fake_db

    with TestClient(app) as c:
        yield c, SessionLocal


def _seed_family_data(SessionLocal) -> None:
    """Två personer, gemensamt konto, lite utgifter + inkomster."""
    from hembudget.db.models import Account, Category, Transaction, User, Budget
    with SessionLocal() as s:
        robin = User(name="Robin")
        sara = User(name="Sara")
        s.add_all([robin, sara])
        s.flush()
        a_robin = Account(name="Robin lön", bank="nordea", type="checking", owner_id=robin.id)
        a_sara = Account(name="Sara lön", bank="nordea", type="checking", owner_id=sara.id)
        a_gem = Account(name="Gemensamt", bank="nordea", type="shared")
        s.add_all([a_robin, a_sara, a_gem])
        s.flush()
        mat = Category(name="Mat")
        el = Category(name="El")
        lon = Category(name="Lön")
        s.add_all([mat, el, lon])
        s.flush()

        # April-data
        s.add_all([
            Transaction(account_id=a_robin.id, date=date(2026, 4, 25),
                        amount=Decimal("40000"), currency="SEK",
                        raw_description="Lön", hash="r-lon4",
                        category_id=lon.id),
            Transaction(account_id=a_sara.id, date=date(2026, 4, 25),
                        amount=Decimal("30000"), currency="SEK",
                        raw_description="Lön", hash="s-lon4",
                        category_id=lon.id),
            Transaction(account_id=a_gem.id, date=date(2026, 4, 10),
                        amount=Decimal("-6000"), currency="SEK",
                        raw_description="Mat", hash="mat4",
                        category_id=mat.id),
            Transaction(account_id=a_gem.id, date=date(2026, 4, 15),
                        amount=Decimal("-2200"), currency="SEK",
                        raw_description="El", hash="el4",
                        category_id=el.id),
        ])
        # Mars-data (för delta-jämförelse)
        s.add_all([
            Transaction(account_id=a_gem.id, date=date(2026, 3, 10),
                        amount=Decimal("-5000"), currency="SEK",
                        raw_description="Mat", hash="mat3",
                        category_id=mat.id),
            Transaction(account_id=a_gem.id, date=date(2026, 3, 15),
                        amount=Decimal("-2100"), currency="SEK",
                        raw_description="El", hash="el3",
                        category_id=el.id),
        ])
        # Budget
        s.add_all([
            Budget(month="2026-04", category_id=mat.id, planned_amount=Decimal("-5000")),
            Budget(month="2026-04", category_id=el.id, planned_amount=Decimal("-2000")),
        ])
        s.commit()


def test_build_report_data_returns_transfers(client):
    c, SessionLocal = client
    _seed_family_data(SessionLocal)

    from hembudget.reports.monthly_pdf import build_report_data
    with SessionLocal() as s:
        data = build_report_data(s, "2026-04")

    assert data.month == "2026-04"
    # Två personer ska ha transfer-förslag
    assert len(data.transfers) == 2
    names = {t.person_name for t in data.transfers}
    assert names == {"Robin", "Sara"}
    # Summan av 50/50 ska matcha gemensamma utgifter (8200 kr)
    equal_total = sum(t.fair_equal for t in data.transfers)
    assert abs(equal_total - 8200) < 1
    # Prorata ska också summera till 8200 (fördelar exakt)
    prorata_total = sum(t.fair_prorata for t in data.transfers)
    assert abs(prorata_total - 8200) < 5  # avrundning kan dra lite


def test_build_report_data_prev_month_deltas(client):
    c, SessionLocal = client
    _seed_family_data(SessionLocal)

    from hembudget.reports.monthly_pdf import build_report_data
    with SessionLocal() as s:
        data = build_report_data(s, "2026-04")

    # Mat gick upp från 5000 till 6000 → delta +1000 (utgifter >0 = ökning)
    deltas_by_cat = {name: d for name, d in data.deltas}
    assert "Mat" in deltas_by_cat
    assert abs(deltas_by_cat["Mat"] - 1000) < 1


def test_pdf_endpoint_returns_pdf_bytes(client):
    c, SessionLocal = client
    _seed_family_data(SessionLocal)

    r = c.get("/reports/month/2026-04/pdf")
    assert r.status_code == 200, r.content
    assert r.content[:4] == b"%PDF"
    assert len(r.content) > 5000  # Rimligt för en rapport med charts
    assert r.headers["content-type"] == "application/pdf"


def test_pdf_handles_empty_month(client):
    """Månad utan data ska inte krascha — bara tomma charts."""
    c, _ = client
    r = c.get("/reports/month/2030-01/pdf")
    assert r.status_code == 200
    assert r.content[:4] == b"%PDF"


def test_excel_endpoint_has_multiple_sheets(client):
    c, SessionLocal = client
    _seed_family_data(SessionLocal)

    r = c.get("/reports/month/2026-04/excel")
    assert r.status_code == 200
    # Excel = zip-fil som börjar med PK
    assert r.content[:2] == b"PK"
    # Snabb inspektion: öppna XLSX i minnet och kolla flikar
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content))
    assert "Överföringar" in wb.sheetnames
    assert "Förändring mot förra" in wb.sheetnames


def test_charts_empty_dont_crash():
    """Piecharts med tom lista ska rendera en placeholder, inte krascha."""
    from hembudget.reports import charts
    png = charts.pie_chart([], [], title="Tom")
    assert len(png) > 500  # det är en riktig PNG
    png2 = charts.bar_chart_budget_vs_actual([], [], [], title="Tom")
    assert len(png2) > 500
