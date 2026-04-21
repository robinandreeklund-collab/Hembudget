"""SEB Kort Excel-export (.xlsx).

Kolumner:  Datum, Bokfört, Specifikation, Ort, Valuta, Utl. belopp, Belopp

Viktigt: SEB Kort rapporterar KORT-kontots saldo, där positivt belopp
= ökad skuld (ditt köp, din kostnad), negativt = inbetalning från dig.
Vi använder konventionen "negativt = utgift från plånboken" i hela
systemet, så vi inverterar tecknet för att matcha Amex och Nordea.
"""
from __future__ import annotations

import io
import logging
from datetime import date, datetime
from decimal import Decimal

from .base import BankParser, RawTransaction

log = logging.getLogger(__name__)


XLSX_MAGIC = b"PK\x03\x04"


def _to_decimal(val) -> Decimal:
    if val is None or val == "":
        return Decimal("0")
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))
    s = str(val).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


def _to_date(val) -> date | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


class SebKortXlsxParser(BankParser):
    bank = "seb_kort"
    name = "SEB Kort (Mastercard Eurobonus) — Excel"

    HEADER_KEYWORDS = {"datum", "specifikation", "belopp"}

    def detect(self, content: bytes) -> bool:
        if not content.startswith(XLSX_MAGIC):
            return False
        try:
            from openpyxl import load_workbook
        except ImportError:
            log.warning("openpyxl unavailable — cannot detect SEB Kort xlsx")
            return False
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception:
            # Fallback: if it's a valid xlsx with magic bytes but we can't
            # fully parse from the sample, accept it as a SEB-like source.
            # parse() will raise a clearer error if the structure is wrong.
            return True
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True, max_row=15):
                cells = {str(c).strip().lower() for c in row if c is not None}
                if self.HEADER_KEYWORDS.issubset(cells):
                    return True
        return False

    def parse(self, content: bytes) -> list[RawTransaction]:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        # Prefer "Transaktioner" sheet, fall back to first
        sheet_name = "Transaktioner" if "Transaktioner" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]

        header_row: list[str] | None = None
        header_idx = -1
        for i, row in enumerate(ws.iter_rows(values_only=True, max_row=15)):
            lowered = [str(c).strip().lower() if c is not None else "" for c in row]
            if self.HEADER_KEYWORDS.issubset({c for c in lowered if c}):
                header_row = lowered
                header_idx = i
                break

        if header_row is None:
            return []

        def idx(name: str) -> int:
            try:
                return header_row.index(name)
            except ValueError:
                return -1

        i_date = idx("datum")
        i_spec = idx("specifikation")
        i_city = idx("ort")
        i_curr = idx("valuta")
        i_foreign = idx("utl. belopp")
        i_amount = idx("belopp")

        out: list[RawTransaction] = []
        row_counter = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i <= header_idx:
                continue
            if not row or all(c is None or c == "" for c in row):
                continue
            d = _to_date(row[i_date]) if i_date >= 0 else None
            if d is None:
                continue
            raw_amount = _to_decimal(row[i_amount]) if i_amount >= 0 else Decimal("0")
            # Invert sign: SEB reports positive for purchases, negative for
            # payments. Our convention: negative = expense.
            amount = -raw_amount

            spec = str(row[i_spec]).strip() if i_spec >= 0 and row[i_spec] else ""
            city = str(row[i_city]).strip() if i_city >= 0 and row[i_city] else ""
            description = f"{spec} [{city}]" if spec and city else (spec or city)

            currency = str(row[i_curr]).strip() if i_curr >= 0 and row[i_curr] else "SEK"
            foreign = row[i_foreign] if i_foreign >= 0 else None

            out.append(
                RawTransaction(
                    date=d,
                    amount=amount,
                    description=description,
                    currency=currency or "SEK",
                    row_index=row_counter,
                    meta={"foreign_amount": str(foreign) if foreign else None},
                )
            )
            row_counter += 1
        return out
