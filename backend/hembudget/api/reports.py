from __future__ import annotations

import io

from fastapi import APIRouter, Depends
from fastapi.responses import Response
from sqlalchemy.orm import Session

from ..budget.monthly import MonthlyBudgetService
from .deps import db, require_auth

router = APIRouter(prefix="/reports", tags=["reports"], dependencies=[Depends(require_auth)])


@router.get("/month/{month}/excel")
def excel_report(month: str, session: Session = Depends(db)) -> Response:
    """Excel-export med flera flikar: Budget, Överföringar, Förra månaden.

    Graceful — returnerar 501 om openpyxl saknas."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return Response(status_code=501, content="openpyxl ej installerat")

    from ..reports.monthly_pdf import build_report_data, _format_month_label
    data = build_report_data(session, month)
    summary = data.summary

    wb = Workbook()

    # --- Budget-fliken (förvald) ---
    ws = wb.active
    ws.title = f"Budget {month}"
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")

    ws["A1"] = f"Hembudget — Månadsrapport {_format_month_label(month)}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")
    ws["A3"] = "Inkomst"
    ws["B3"] = float(summary.income)
    ws["A4"] = "Utgifter"
    ws["B4"] = float(summary.expenses)
    ws["A5"] = "Sparat"
    ws["B5"] = float(summary.savings)
    ws["A6"] = "Sparkvot"
    ws["B6"] = summary.savings_rate
    for r in range(3, 7):
        ws[f"A{r}"].font = bold

    # Tabellen börjar på rad 8
    header_row = 8
    ws.cell(row=header_row, column=1, value="Grupp")
    ws.cell(row=header_row, column=2, value="Kategori")
    ws.cell(row=header_row, column=3, value="Budgeterat")
    ws.cell(row=header_row, column=4, value="Faktiskt")
    ws.cell(row=header_row, column=5, value="Diff")
    ws.cell(row=header_row, column=6, value="Progress %")
    for col in range(1, 7):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    r = header_row + 1
    lines_by_group: dict[str, list] = {}
    for l in summary.lines:
        key = "Inkomster" if l.kind == "income" else (l.group or "Övrigt")
        lines_by_group.setdefault(key, []).append(l)
    for group, group_lines in sorted(
        lines_by_group.items(), key=lambda kv: (kv[0] == "Inkomster", kv[0])
    ):
        for l in group_lines:
            ws.cell(row=r, column=1, value=group)
            ws.cell(row=r, column=2, value=l.category)
            ws.cell(row=r, column=3, value=float(l.planned))
            ws.cell(row=r, column=4, value=float(l.actual))
            ws.cell(row=r, column=5, value=float(l.diff))
            ws.cell(
                row=r, column=6,
                value=(l.progress_pct / 100 if l.planned != 0 else None),
            )
            ws.cell(row=r, column=6).number_format = "0 %"
            r += 1

    # --- Överföringar ---
    if data.transfers:
        ws2 = wb.create_sheet("Överföringar")
        headers = [
            "Person", "Inkomst", "Andel", "50/50-split",
            "Prorata", "Redan betalt",
        ]
        for i, h in enumerate(headers, start=1):
            c = ws2.cell(row=1, column=i, value=h)
            c.font = header_font
            c.fill = header_fill
        for i, t in enumerate(data.transfers, start=2):
            ws2.cell(row=i, column=1, value=t.person_name)
            ws2.cell(row=i, column=2, value=t.income)
            ws2.cell(row=i, column=3, value=t.income_share_pct / 100).number_format = "0 %"
            ws2.cell(row=i, column=4, value=t.fair_equal)
            ws2.cell(row=i, column=5, value=t.fair_prorata)
            ws2.cell(row=i, column=6, value=t.already_paid)

    # --- Förändring mot föregående månad ---
    if data.deltas:
        ws3 = wb.create_sheet("Förändring mot förra")
        ws3.cell(row=1, column=1, value="Kategori").font = header_font
        ws3.cell(row=1, column=1).fill = header_fill
        ws3.cell(row=1, column=2, value="Diff").font = header_font
        ws3.cell(row=1, column=2).fill = header_fill
        for i, (name, d) in enumerate(data.deltas, start=2):
            ws3.cell(row=i, column=1, value=name)
            ws3.cell(row=i, column=2, value=d)

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="hembudget-{month}.xlsx"'},
    )


@router.get("/month/{month}/pdf")
def pdf_report(month: str, session: Session = Depends(db)) -> Response:
    """Rik månadsrapport som PDF: KPI-ruta, piecharts, transfer-förslag,
    budget vs utfall, förändring mot förra månaden, grupperad tabell.

    Graceful — returnerar 501 om reportlab eller matplotlib saknas."""
    try:
        from ..reports.monthly_pdf import build_report_data, render_pdf
    except ImportError as exc:
        return Response(
            status_code=501,
            content=f"PDF-rapport kräver reportlab + matplotlib: {exc}",
        )

    data = build_report_data(session, month)
    pdf_bytes = render_pdf(data)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="hembudget-{month}.pdf"'},
    )
